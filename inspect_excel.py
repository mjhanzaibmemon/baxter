#!/usr/bin/env python3
"""
Inspect Excel files to see their structure and sample data
"""
import openpyxl
from pathlib import Path

def inspect_excel(filepath):
    print(f"\n{'='*70}")
    print(f"FILE: {filepath.name}")
    print(f"Size: {filepath.stat().st_size:,} bytes")
    print(f"{'='*70}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"\nSheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        print(f"Columns ({len(headers)}): {headers[:10]}...")  # First 10 columns
        
        # Count rows
        total_rows = sum(1 for _ in sheet.iter_rows(min_row=2))
        print(f"Total data rows: {total_rows}")
        
        # Show first 3 data rows
        print("\nFirst 3 rows:")
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True), 1):
            print(f"  Row {i}: {row[:5]}...")  # First 5 columns
    
    wb.close()

# Check the files that were sent
sample_dir = Path("sample_data")

print("\n" + "="*70)
print("INSPECTING EXCEL FILES FROM EMAIL ATTACHMENTS")
print("="*70)

# MS_Kargo file
ms_kargo = sample_dir / "MS_Kargo.xlsx"
if ms_kargo.exists():
    inspect_excel(ms_kargo)

# CS_RMA_DETAIL file  
cs_rma = sample_dir / "CS_RMA_DETAIL_Embed_Excel_061073437323.xlsx"
if cs_rma.exists():
    inspect_excel(cs_rma)

print("\n" + "="*70)
print("INSPECTION COMPLETE")
print("="*70)
