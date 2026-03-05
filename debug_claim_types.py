#!/usr/bin/env python3
"""Debug script to check what claim_type values exist in CS_RMA files"""
import openpyxl
import sys
from collections import Counter

if len(sys.argv) < 2:
    print("Usage: python debug_claim_types.py path/to/CS_RMA_file.xlsx")
    sys.exit(1)

filepath = sys.argv[1]

wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

if 'Data' not in wb.sheetnames:
    print("ERROR: 'Data' sheet not found")
    sys.exit(1)

ws = wb['Data']

# Find header row
header_row_idx = None
header_cells = None
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
    row_vals = [c.value for c in row]
    row_strs = [str(v).strip().lower() if v else "" for v in row_vals]
    if any("effective date" in s for s in row_strs):
        header_row_idx = row_idx
        header_cells = row_vals
        break

if header_row_idx is None:
    print("ERROR: Could not find header row")
    sys.exit(1)

print(f"Header row found at row {header_row_idx}")
print(f"\nAll headers:")
for i, h in enumerate(header_cells):
    if h:
        print(f"  Col {i}: {h}")

# Find "Returned Reason" columns
reason_cols = []
for i, h in enumerate(header_cells):
    if h and "returned reason" in str(h).lower():
        reason_cols.append((i, h))

print(f"\n'Returned Reason' columns found: {len(reason_cols)}")
for col_idx, col_name in reason_cols:
    print(f"  Col {col_idx}: {col_name}")

# Count unique values in first "Returned Reason" column
if reason_cols:
    col_idx = reason_cols[0][0]
    values = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if col_idx < len(row):
            val = row[col_idx]
            if val is not None:
                values.append(str(val).strip())
            else:
                values.append(None)
    
    counter = Counter(values)
    print(f"\nUnique values in '{reason_cols[0][1]}' (Total rows: {len(values)}):")
    for val, count in counter.most_common(20):
        print(f"  {val!r}: {count}")
    
    print(f"\nTotal unique values: {len(counter)}")
    print(f"NULL/None values: {counter.get(None, 0)}")

wb.close()
