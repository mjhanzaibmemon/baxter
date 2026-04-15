"""
excel_parser.py — Parse .xlsx and .csv attachments into list of dicts.

Supports:
    - MS_Kargo Excel schema  (APPOINTMENT_DATE, PRO_NUMBER, SSCC18, SCAC, ORDER_ID, SHIPMENT_ID)
    - Result_*.csv shipments (order_number, sscc18, scac, kargo_shipment_id, appointment_date)
    - CS_RMA_DETAIL          (Summary sheet: Carrier/BP, Credit, Count)
    - CS_RMA_DETAIL          (Data sheet: individual claim rows → claim_details)
"""
import csv
import io
import logging
from datetime import datetime
import openpyxl

logger = logging.getLogger(__name__)


def _parse_datetime_flexible(value):
    """Parse datetimes from Excel or CSV values, returning None on failure."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in (
        None,
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %I:%M %p",
        "%Y-%m-%d",
    ):
        try:
            if fmt is None:
                return datetime.fromisoformat(text)
            return datetime.strptime(text, fmt)
        except Exception:
            continue

    logger.warning(f"Could not parse date: {value}")
    return None


def _normalize_header(h) -> str:
    """Lowercase + strip a header string for comparison."""
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_").replace("/", "_")


def parse_ms_kargo(file_bytes: bytes) -> list[dict]:
    """
    Parse the MS_Kargo Excel format.
    Expected headers (row 1): APPOINTMENT_DATE, PRO_NUMBER, SSCC18, SCAC, ORDER_ID, SHIPMENT_ID
    Returns list of dicts ready for DB insert.
    """
    if not file_bytes or len(file_bytes) < 100:
        logger.warning("MS_Kargo: Empty or invalid file")
        return []
    
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"MS_Kargo: Failed to open Excel file: {e}")
        return []

    # Try 'Sheet1' first, fallback to active sheet
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    headers = [_normalize_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.debug(f"MS_Kargo headers detected: {headers}")

    field_map = {
        "appointment_date": "appointment_date",
        "pro_number":       "pro_number",
        "sscc18":           "sscc18",
        "scac":             "scac",
        "order_id":         "order_id",
        "shipment_id":      "shipment_id",
    }

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        record = {}
        for col_idx, raw_header in enumerate(headers):
            if col_idx >= len(row):
                break
            db_field = field_map.get(raw_header)
            if db_field:
                val = row[col_idx]
                # Ensure appointment_date is datetime
                if db_field == "appointment_date":
                    record[db_field] = _parse_datetime_flexible(val)
                else:
                    record[db_field] = str(val).strip() if val is not None else None

        # Only include rows with minimum required fields
        if record.get("sscc18") and record.get("appointment_date"):
            rows.append(record)

    logger.info(f"Parsed {len(rows)} valid rows from MS_Kargo sheet")
    wb.close()
    return rows


def parse_result_csv(file_bytes: bytes) -> list[dict]:
    """
    Parse Paul's Result_27.csv shipment export.
    Expected headers: order_number, sscc18, scac, kargo_shipment_id, appointment_date
    Returns list of dicts ready for shipments insert.
    """
    if not file_bytes:
        logger.warning("Result CSV: Empty or invalid file")
        return []

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        logger.warning("Result CSV: Missing header row")
        return []

    headers = {_normalize_header(name): name for name in reader.fieldnames if name}
    required = {"order_number", "sscc18", "kargo_shipment_id", "appointment_date"}
    if not required.issubset(headers.keys()):
        logger.warning(f"Result CSV: Missing required columns. Found: {sorted(headers.keys())}")
        return []

    rows = []
    skipped_missing_date = 0
    skipped_missing_sscc = 0

    for raw_row in reader:
        appointment_date = _parse_datetime_flexible(raw_row.get(headers["appointment_date"]))
        sscc18 = (raw_row.get(headers["sscc18"]) or "").strip()

        if not sscc18:
            skipped_missing_sscc += 1
            continue
        if appointment_date is None:
            skipped_missing_date += 1
            continue

        rows.append({
            "appointment_date": appointment_date,
            "pro_number": None,
            "sscc18": sscc18,
            "scac": (raw_row.get(headers.get("scac")) or "").strip() or None,
            "order_id": (raw_row.get(headers["order_number"]) or "").strip() or None,
            "shipment_id": (raw_row.get(headers["kargo_shipment_id"]) or "").strip() or None,
        })

    logger.info(
        "Parsed %s valid shipment rows from Result CSV (%s skipped missing appointment_date, %s skipped missing sscc18)",
        len(rows),
        skipped_missing_date,
        skipped_missing_sscc,
    )
    return rows


def parse_rma_detail(file_bytes: bytes) -> list[dict]:
    """
    Parse the CS_RMA_DETAIL Excel — Summary sheet only.
    Headers start at row 3: Carrier/BP, Credit, Count
    Returns list of dicts for rma_credits table.
    """
    if not file_bytes or len(file_bytes) < 100:
        logger.warning("CS_RMA_DETAIL: Empty or invalid file")
        return []
    
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"CS_RMA_DETAIL: Failed to open Excel file: {e}")
        return []

    ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.active

    rows = []
    header_row_idx = None

    # Find the header row dynamically
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_lower = [str(v).strip().lower() if v else "" for v in row]
        if "carrier/bp" in row_lower or "carrier_bp" in row_lower:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        logger.warning("CS_RMA_DETAIL: Could not find header row in Summary sheet")
        return rows

    # Read data rows after header
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue
        carrier_bp, credit, count = row[0], row[1], row[2]
        if carrier_bp is None:
            carrier_bp = 'Unknown'
        rows.append({
            "carrier_bp": str(carrier_bp).strip(),
            "credit":     float(credit) if credit is not None else None,
            "count":      int(count) if count is not None else None,
        })

    logger.info(f"Parsed {len(rows)} RMA credit rows")
    wb.close()
    return rows


def parse_all_orders_csv(file_bytes: bytes) -> list[dict]:
    """
    Parse Paul's AllOrders.csv — order status data.
    Columns: Order_Number, Order_Status, Missed_LPNs
    Returns list of dicts with order_number, order_status, missed_lpns.
    """
    if not file_bytes:
        logger.warning("AllOrders CSV: Empty or invalid file")
        return []

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        logger.warning("AllOrders CSV: Missing header row")
        return []

    headers = {_normalize_header(name): name for name in reader.fieldnames if name}
    required = {"order_number", "order_status"}
    if not required.issubset(headers.keys()):
        logger.warning(f"AllOrders CSV: Missing required columns. Found: {sorted(headers.keys())}")
        return []

    rows = []
    for raw_row in reader:
        order_number = (raw_row.get(headers["order_number"]) or "").strip()
        order_status = (raw_row.get(headers["order_status"]) or "").strip()
        missed_lpns = (raw_row.get(headers.get("missed_lpns", ""), "") or "").strip()

        if not order_number or not order_status:
            continue

        rows.append({
            "order_number": order_number,
            "order_status": order_status,
            "missed_lpns": missed_lpns,
        })

    logger.info(f"Parsed {len(rows)} rows from AllOrders CSV")
    return rows


def parse_order_level_csv(file_bytes: bytes) -> list[dict]:
    """
    Parse Kargo order_level CSV files.
    Columns: orderNumber, orderStatus, missedLpns
    orderNumber is a 16-digit Kargo number (may have leading tab/whitespace).
    First 8 digits = shipment prefix, last 8 = original order number.
    orderStatus: COMPLETED -> Perfect, INCOMPLETE -> Short.
    missedLpns: pipe-delimited SSCC18 values (for INCOMPLETE orders).
    Returns list of dicts for update_kargo_order_status().
    """
    if not file_bytes:
        logger.warning("Order level CSV: Empty or invalid file")
        return []

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        logger.warning("Order level CSV: Missing header row")
        return []

    headers = {_normalize_header(name): name for name in reader.fieldnames if name}
    required = {"ordernumber", "orderstatus"}
    if not required.issubset(headers.keys()):
        logger.warning(f"Order level CSV: Missing required columns. Found: {sorted(headers.keys())}")
        return []

    STATUS_MAP = {"COMPLETED": "Perfect", "INCOMPLETE": "Short"}

    rows = []
    for raw_row in reader:
        raw_order = (raw_row.get(headers["ordernumber"]) or "").strip()
        raw_status = (raw_row.get(headers["orderstatus"]) or "").strip().upper()
        missed_lpns = (raw_row.get(headers.get("missedlpns", ""), "") or "").strip()

        if not raw_order or not raw_status:
            continue

        # Strip non-digit chars and extract original order (last 8 digits)
        digits_only = "".join(c for c in raw_order if c.isdigit())
        if len(digits_only) < 8:
            logger.warning(f"Order level CSV: orderNumber too short: '{raw_order}'")
            continue

        original_order = digits_only[8:] if len(digits_only) > 8 else digits_only
        shipment_id = digits_only[:8] if len(digits_only) > 8 else ""
        mapped_status = STATUS_MAP.get(raw_status, raw_status)

        # Parse pipe-delimited SSCC18 list
        sscc18_list = [s.strip() for s in missed_lpns.split("|") if s.strip()] if missed_lpns else []

        rows.append({
            "order_number": original_order,
            "order_status": mapped_status,
            "kargo_order_number": digits_only,
            "shipment_id": shipment_id,
            "missed_lpns": sscc18_list,
        })

    logger.info(f"Parsed {len(rows)} rows from order_level CSV")
    return rows


def detect_and_parse(filename: str, file_bytes: bytes):
    """
    Auto-detect file type by filename and parse accordingly.
    Returns (type, rows) where type is 'shipments', 'rma', or 'unknown'.
    For CS_RMA files, returns (type, rows, claim_detail_rows).
    """
    fname_lower = filename.lower()
    if "allorders" in fname_lower or "all_orders" in fname_lower:
        return "order_status", parse_all_orders_csv(file_bytes), []
    elif "order_level" in fname_lower:
        return "order_level", parse_order_level_csv(file_bytes), []
    elif "ms_kargo" in fname_lower or "kargo" in fname_lower:
        return "shipments", parse_ms_kargo(file_bytes), []
    elif fname_lower.endswith(".csv") and ("result" in fname_lower or "shipment" in fname_lower):
        return "shipments", parse_result_csv(file_bytes), []
    elif "cs_rma" in fname_lower or "rma" in fname_lower:
        rma_rows = parse_rma_detail(file_bytes)
        claim_rows = parse_rma_data_sheet(file_bytes)
        return "rma", rma_rows, claim_rows
    else:
        if fname_lower.endswith(".csv"):
            logger.info(f"Unknown CSV file type '{filename}', attempting Result CSV parse")
            try:
                rows = parse_result_csv(file_bytes)
                return "shipments", rows, []
            except Exception:
                logger.warning(f"Could not parse '{filename}' as Result CSV either — skipping")
                return "unknown", [], []

        # Try MS_Kargo format as default (check headers)
        logger.info(f"Unknown file type '{filename}', attempting MS_Kargo parse")
        try:
            rows = parse_ms_kargo(file_bytes)
            return "shipments", rows, []
        except Exception:
            logger.warning(f"Could not parse '{filename}' as MS_Kargo either — skipping")
            return "unknown", [], []


def parse_rma_data_sheet(file_bytes: bytes) -> list[dict]:
    """
    Parse the CS_RMA_DETAIL Excel — 'Data' sheet with individual claim rows.
    Headers are in row 5. Data starts after header row.
    Returns list of dicts for claim_details table.

    NOTE: Different CS_RMA files have DIFFERENT column layouts:
      - Some have 35 columns (full detail)
      - Some have only 13 columns (compact)
    
    We use dynamic header matching by scanning ALL column names, not a fixed
    dict (which breaks when duplicate header names exist like "Returned Reason").
    """
    if not file_bytes or len(file_bytes) < 100:
        logger.warning("CS_RMA_DETAIL Data: Empty or invalid file")
        return []
    
    try:
        wb = openpyxl.load_workbook(
            filename=io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        logger.error(f"CS_RMA_DETAIL Data: Failed to open Excel file: {e}")
        return []

    if "Data" not in wb.sheetnames:
        logger.warning("CS_RMA_DETAIL: 'Data' sheet not found — no claim details to parse")
        return []

    ws = wb["Data"]

    # Find header row dynamically (contains "Effective Date")
    header_row_idx = None
    header_cells = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
        row_vals = [c.value for c in row]
        row_strs = [str(v).strip().lower() if v else "" for v in row_vals]
        if any("effective date" in s for s in row_strs):
            header_row_idx = row_idx
            header_cells = row_vals
            break

    if header_row_idx is None:
        logger.warning("CS_RMA_DETAIL Data sheet: Could not find header row")
        return []

    # Build column index list: (index, lowercase_header_name)
    # DO NOT use a dict — duplicate header names (e.g., "Returned Reason") will overwrite
    all_headers = []
    for i, h in enumerate(header_cells):
        if h is not None:
            all_headers.append((i, str(h).strip().lower()))

    logger.info(f"Data sheet has {len(all_headers)} header columns at row {header_row_idx}")

    # Helper to find first column matching a keyword
    def find_col(keyword):
        for idx, name in all_headers:
            if keyword in name:
                return idx
        return None

    # Helper to find ALL columns matching a keyword (returns list of indices)
    def find_all_cols(keyword):
        return [idx for idx, name in all_headers if keyword in name]

    # Find required columns
    col_date = find_col("effective date")
    col_credit = find_col("credit issued")

    # Find optional columns (may not exist in compact 13-col files)
    col_order = find_col("original order number")
    col_shipment = find_col("shipment number")
    col_carrier = find_col("concatenation carrier number")

    # Returned Reason appears TWICE in 35-col files:
    #   First  = internal code  (K12, K7, etc.)
    #   Second = readable type  (SHORTAGE, DAMAGE, etc.)
    # In 13-col files, there may still be two cols.
    # We always want the LAST "Returned Reason" column (readable type).
    reason_cols = find_all_cols("returned reason")
    col_reason = reason_cols[-1] if reason_cols else None

    # Schema v2: additional columns from 35-col files
    col_rma_status = find_col("rma status")
    col_rma_date = find_col("effective date")  # same col, but we also look for rma_date below
    do_ty_cols = find_all_cols("do ty")
    col_doc_type = do_ty_cols[0] if do_ty_cols else None
    col_return_doc_type = do_ty_cols[-1] if len(do_ty_cols) > 1 else None
    col_rma_order = find_col("rma order")
    col_po = find_col("po #") or find_col("po")
    col_order_type = find_col("or ty")
    col_reference_number_qualifier = find_col("reference number qualifier")
    col_bol_number = find_col("bol number")
    col_original_line_number = find_col("original line number")
    col_returned_material_status = find_col("returned material status")
    col_contact_name = find_col("contact name")
    col_description = find_col("description")
    col_line_number = None
    for idx, name in all_headers:
        if "line number" in name and "original" not in name:
            col_line_number = idx
            break
    address_cols = find_all_cols("address number")
    col_address_number = address_cols[0] if address_cols else None
    col_address_name = address_cols[-1] if len(address_cols) > 1 else None
    col_ship_to_number = find_col("ship to number")
    # "Ship To" appears twice: "Ship To Number" (numeric ID) and "Ship To" (name).
    # We want the NAME column, which is the shorter header without "number".
    ship_to_cols = find_all_cols("ship to")
    col_ship_to = None
    for idx, name in all_headers:
        if idx in ship_to_cols and "number" not in name:
            col_ship_to = idx
            break
    if col_ship_to is None and ship_to_cols:
        col_ship_to = ship_to_cols[-1]  # fallback to last
    col_item = find_col("2nd item") or find_col("item number")
    col_unit_of_measure = find_col("um")
    col_quantity = find_col("quantity")
    col_reason_code = reason_cols[0] if len(reason_cols) > 1 else None
    # Branch appears twice (code + name).
    # We store both branch code and readable branch name.
    branch_cols = find_all_cols("branch")
    col_branch_code = branch_cols[0] if branch_cols else None
    col_branch = branch_cols[-1] if branch_cols else None
    business_unit_cols = find_all_cols("business unit")
    col_business_unit_code = business_unit_cols[0] if business_unit_cols else None
    col_business_unit = business_unit_cols[-1] if business_unit_cols else None
    col_serial_number_lot = find_col("serial number lot")
    col_lot_serial_number = find_col("lot serial number")

    if col_date is None or col_credit is None:
        logger.warning(
            f"CS_RMA_DETAIL Data sheet: Missing required columns "
            f"(date={col_date}, credit={col_credit})"
        )
        return []

    logger.info(
        f"Data sheet column mapping: date={col_date}, order={col_order}, "
        f"shipment={col_shipment}, credit={col_credit}, carrier={col_carrier}, "
        f"reason={col_reason} (from {len(reason_cols)} 'Returned Reason' cols)"
    )
    logger.info(
        f"Schema v2 columns: rma_status={col_rma_status}, doc_type={col_doc_type}, "
        f"rma_order={col_rma_order}, po={col_po}, order_type={col_order_type}, "
        f"ship_to={col_ship_to}, item={col_item}, quantity={col_quantity}, "
        f"reason_code={col_reason_code}, branch={col_branch}"
    )
    logger.info(
        f"Schema v3 columns: ref_qualifier={col_reference_number_qualifier}, bol={col_bol_number}, "
        f"orig_line={col_original_line_number}, line={col_line_number}, returned_material_status={col_returned_material_status}, "
        f"contact_name={col_contact_name}, description={col_description}, address_number={col_address_number}, "
        f"address_name={col_address_name}, ship_to_number={col_ship_to_number}, um={col_unit_of_measure}, "
        f"branch_code={col_branch_code}, business_unit_code={col_business_unit_code}, business_unit={col_business_unit}, "
        f"serial_number_lot={col_serial_number_lot}, lot_serial_number={col_lot_serial_number}, return_doc_type={col_return_doc_type}"
    )

    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Skip empty rows
        if all(v is None for v in row):
            continue

        row_len = len(row)

        # Get effective date
        date_val = row[col_date] if col_date < row_len else None
        if date_val is None:
            skipped += 1
            continue

        # Parse date
        if isinstance(date_val, datetime):
            claim_date = date_val
        else:
            try:
                claim_date = datetime.fromisoformat(str(date_val))
            except Exception:
                skipped += 1
                continue

        # Get credit amount
        credit_val = row[col_credit] if col_credit < row_len else None
        if credit_val is None:
            skipped += 1
            continue
        try:
            claim_amount = float(credit_val)
        except (ValueError, TypeError):
            skipped += 1
            continue

        # Get optional fields safely
        def safe_str(val):
            if val is None:
                return None
            s = str(val).strip()
            return s if s else None

        def safe_int_str(val):
            if val is None:
                return None
            s = str(val).strip()
            if not s:
                return None
            try:
                return str(int(float(s)))
            except (ValueError, TypeError):
                return s

        order_id = safe_int_str(row[col_order]) if col_order is not None and col_order < row_len else None
        sscc18 = safe_int_str(row[col_shipment]) if col_shipment is not None and col_shipment < row_len else None
        carrier_bp = safe_str(row[col_carrier]) if col_carrier is not None and col_carrier < row_len else None
        carrier_bp = carrier_bp or 'Unknown'  # Prevent NULL carrier_bp in DB
        claim_type_raw = safe_str(row[col_reason]) if col_reason is not None and col_reason < row_len else None

        # Schema v2: extract additional fields
        rma_status = safe_str(row[col_rma_status]) if col_rma_status is not None and col_rma_status < row_len else None
        doc_type_val = safe_str(row[col_doc_type]) if col_doc_type is not None and col_doc_type < row_len else None
        return_doc_type_val = safe_str(row[col_return_doc_type]) if col_return_doc_type is not None and col_return_doc_type < row_len else None
        rma_order_val = safe_int_str(row[col_rma_order]) if col_rma_order is not None and col_rma_order < row_len else None
        po_val = safe_str(row[col_po]) if col_po is not None and col_po < row_len else None
        order_type_val = safe_str(row[col_order_type]) if col_order_type is not None and col_order_type < row_len else None
        reference_number_qualifier_val = safe_str(row[col_reference_number_qualifier]) if col_reference_number_qualifier is not None and col_reference_number_qualifier < row_len else None
        bol_number_val = safe_str(row[col_bol_number]) if col_bol_number is not None and col_bol_number < row_len else None
        returned_material_status_val = safe_str(row[col_returned_material_status]) if col_returned_material_status is not None and col_returned_material_status < row_len else None
        contact_name_val = safe_str(row[col_contact_name]) if col_contact_name is not None and col_contact_name < row_len else None
        description_val = safe_str(row[col_description]) if col_description is not None and col_description < row_len else None
        address_number_val = safe_int_str(row[col_address_number]) if col_address_number is not None and col_address_number < row_len else None
        address_name_val = safe_str(row[col_address_name]) if col_address_name is not None and col_address_name < row_len else None
        ship_to_number_val = safe_int_str(row[col_ship_to_number]) if col_ship_to_number is not None and col_ship_to_number < row_len else None
        ship_to_val = safe_str(row[col_ship_to]) if col_ship_to is not None and col_ship_to < row_len else None
        item_val = safe_str(row[col_item]) if col_item is not None and col_item < row_len else None
        unit_of_measure_val = safe_str(row[col_unit_of_measure]) if col_unit_of_measure is not None and col_unit_of_measure < row_len else None
        reason_code_val = safe_str(row[col_reason_code]) if col_reason_code is not None and col_reason_code < row_len else None
        branch_code_val = safe_str(row[col_branch_code]) if col_branch_code is not None and col_branch_code < row_len else None
        branch_val = safe_str(row[col_branch]) if col_branch is not None and col_branch < row_len else None
        business_unit_code_val = safe_str(row[col_business_unit_code]) if col_business_unit_code is not None and col_business_unit_code < row_len else None
        business_unit_val = safe_str(row[col_business_unit]) if col_business_unit is not None and col_business_unit < row_len else None
        serial_number_lot_val = safe_str(row[col_serial_number_lot]) if col_serial_number_lot is not None and col_serial_number_lot < row_len else None
        lot_serial_number_val = safe_str(row[col_lot_serial_number]) if col_lot_serial_number is not None and col_lot_serial_number < row_len else None

        original_line_number_val = None
        if col_original_line_number is not None and col_original_line_number < row_len and row[col_original_line_number] is not None:
            try:
                original_line_number_val = int(float(str(row[col_original_line_number]).strip()))
            except (ValueError, TypeError):
                original_line_number_val = None

        line_number_val = None
        if col_line_number is not None and col_line_number < row_len and row[col_line_number] is not None:
            try:
                line_number_val = int(float(str(row[col_line_number]).strip()))
            except (ValueError, TypeError):
                line_number_val = None

        # Parse quantity (integer)
        quantity_val = None
        if col_quantity is not None and col_quantity < row_len and row[col_quantity] is not None:
            try:
                quantity_val = int(float(str(row[col_quantity]).strip()))
            except (ValueError, TypeError):
                quantity_val = None

        # Parse RMA date (Effective Date is claim_date, but we store it as rma_date too)
        rma_date_val = claim_date  # same date column for now

        # Normalize claim_type for consistent Grafana filtering:
        # "OBSOLETE - SHORTAGE - CC" → "SHORTAGE"
        # "OBSOLETE - DAMAGE -CC" → "DAMAGE"
        # "DAMAGED IN TRANSIT FROM BAXTER" → "DAMAGE"
        # "DAMAGED RETURN-CUSTOMER" → "DAMAGE"
        if claim_type_raw:
            upper = claim_type_raw.upper()
            if "SHORTAGE" in upper or "SHORT" in upper:
                claim_type = "SHORTAGE"
            elif "DAMAGE" in upper:
                claim_type = "DAMAGE"
            else:
                claim_type = claim_type_raw
        else:
            claim_type = "SHORTAGE"

        rows.append({
            "claim_date": claim_date,
            "order_id": order_id,
            "sscc18": sscc18,
            "claim_amount": claim_amount,
            "claim_type": claim_type,
            "carrier_bp": carrier_bp,
            # Schema v2 fields
            "rma_status": rma_status,
            "rma_date": rma_date_val,
            "doc_type": doc_type_val,
            "return_doc_type": return_doc_type_val,
            "rma_order_number": rma_order_val,
            "po_number": po_val,
            "order_type": order_type_val,
            "reference_number_qualifier": reference_number_qualifier_val,
            "bol_number": bol_number_val,
            "original_line_number": original_line_number_val,
            "line_number": line_number_val,
            "returned_material_status": returned_material_status_val,
            "contact_name": contact_name_val,
            "description": description_val,
            "address_number": address_number_val,
            "address_name": address_name_val,
            "ship_to_number": ship_to_number_val,
            "ship_to_name": ship_to_val,
            "item_number": item_val,
            "unit_of_measure": unit_of_measure_val,
            "quantity": quantity_val,
            "reason_code": reason_code_val,
            "reason_text": claim_type_raw,  # original un-normalized reason text
            "branch_code": branch_code_val,
            "branch": branch_val,
            "business_unit_code": business_unit_code_val,
            "business_unit": business_unit_val,
            "serial_number_lot": serial_number_lot_val,
            "lot_serial_number": lot_serial_number_val,
        })

    logger.info(f"Parsed {len(rows)} claim detail rows from Data sheet (skipped {skipped})")
    wb.close()
    return rows
